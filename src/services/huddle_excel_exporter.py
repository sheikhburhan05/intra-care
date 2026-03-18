"""Export huddle analysis JSON files for one or more patients into data.xlsx."""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Sheet definitions ─────────────────────────────────────────────────────────

SHEET_MEDICATION = "medication_to_diagnosis"
SHEET_LAB = "lab_report_to_diagnosis"
SHEET_HUDDLE = "huddle_summary_notes"

HEADERS: dict[str, list[str]] = {
    SHEET_MEDICATION: [
        "patient_id",
        "medication",
        "implied_condition",
        "icd10_code",
        "evidence",
    ],
    SHEET_LAB: [
        "patient_id",
        "lab_report_id",
        "lab_analyte",
        "lab_value",
        "expected_value",
        "implied_condition",
        "icd10_code",
        "evidence",
    ],
    SHEET_HUDDLE: [
        "patient_id",
        "medication_to_diagnosis_summary",
        "lab_report_to_diagnosis_summary",
        "actions",
    ],
}

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Calibri", size=10)
ALT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

# Column widths per sheet (same order as HEADERS)
COL_WIDTHS: dict[str, list[int]] = {
    SHEET_MEDICATION: [14, 30, 30, 12, 60],
    SHEET_LAB: [14, 40, 28, 18, 22, 30, 12, 60],
    SHEET_HUDDLE: [14, 50, 50, 80],
}


class HuddleExcelExporter:
    """Append huddle analysis data for multiple patients into data.xlsx."""

    def __init__(self, output_path: str | Path = "data.xlsx") -> None:
        self.output_path = Path(output_path)

    def export(self, patient_ids: list[str], json_dir: str | Path = ".") -> Path:
        """Read <patient_id>.json for each ID and write rows to data.xlsx.

        If the file exists the rows are appended; otherwise a new workbook is created.
        Returns the path of the written file.
        """
        json_dir = Path(json_dir)

        if self.output_path.exists():
            wb = openpyxl.load_workbook(self.output_path)
        else:
            wb = Workbook()
            wb.remove(wb.active)  # remove default empty sheet

        sheets = self._ensure_sheets(wb)

        loaded = 0
        skipped: list[str] = []

        for pid in patient_ids:
            pid = str(pid).strip()
            if not pid:
                continue
            json_path = json_dir / f"{pid}.json"
            if not json_path.exists():
                print(f"  [skip] {json_path} not found.")
                skipped.append(pid)
                continue

            try:
                analysis = json.loads(json_path.read_text(encoding="utf-8"))
            except (OSError, ValueError) as exc:
                print(f"  [skip] Could not read {json_path}: {exc}")
                skipped.append(pid)
                continue

            self._write_medication_rows(sheets[SHEET_MEDICATION], pid, analysis)
            self._write_lab_rows(sheets[SHEET_LAB], pid, analysis)
            self._write_huddle_row(sheets[SHEET_HUDDLE], pid, analysis)
            loaded += 1
            print(f"  [ok]   {pid}")

        for name, ws in sheets.items():
            self._apply_styles(ws, name)

        wb.save(self.output_path)
        print(f"\nSaved {loaded} patient(s) to {self.output_path.resolve()}")
        if skipped:
            print(f"Skipped: {', '.join(skipped)}")
        return self.output_path

    # ── Sheet bootstrap ───────────────────────────────────────────────────────

    def _ensure_sheets(self, wb: Workbook) -> dict[str, Any]:
        sheets: dict[str, Any] = {}
        for name in (SHEET_MEDICATION, SHEET_LAB, SHEET_HUDDLE):
            if name in wb.sheetnames:
                ws = wb[name]
            else:
                ws = wb.create_sheet(name)
                self._write_header(ws, HEADERS[name])
            sheets[name] = ws
        return sheets

    @staticmethod
    def _write_header(ws: Any, headers: list[str]) -> None:
        ws.append(headers)
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Row writers ───────────────────────────────────────────────────────────

    @staticmethod
    def _write_medication_rows(ws: Any, patient_id: str, analysis: dict[str, Any]) -> None:
        for gap in analysis.get("medication_to_diagnosis", {}).get("suspected_gaps", []):
            ws.append([
                patient_id,
                gap.get("medication", ""),
                gap.get("implied_condition", ""),
                gap.get("icd10_code", ""),
                gap.get("evidence", ""),
            ])

    @staticmethod
    def _write_lab_rows(ws: Any, patient_id: str, analysis: dict[str, Any]) -> None:
        for gap in analysis.get("lab_report_to_diagnosis", {}).get("suspected_gaps", []):
            ws.append([
                patient_id,
                gap.get("lab_report_id", ""),
                gap.get("lab_analyte", ""),
                gap.get("lab_value", ""),
                gap.get("expected_value", ""),
                gap.get("implied_condition", ""),
                gap.get("icd10_code", ""),
                gap.get("evidence", ""),
            ])

        # Combined multi-report gaps (contributing_report_ids joined as lab_report_id)
        for gap in analysis.get("combined_lab_report_to_diagnosis", {}).get("suspected_gaps", []):
            report_ids = ", ".join(gap.get("contributing_report_ids", []))
            ws.append([
                patient_id,
                report_ids,
                "",  # no single analyte
                "",  # no single value
                "",  # no single expected range
                gap.get("implied_condition", ""),
                gap.get("icd10_code", ""),
                gap.get("evidence", ""),
            ])

    @staticmethod
    def _write_huddle_row(ws: Any, patient_id: str, analysis: dict[str, Any]) -> None:
        med_summary = analysis.get("medication_to_diagnosis", {}).get("summary", "")
        lab_summary = analysis.get("lab_report_to_diagnosis", {}).get("narrative_summary", "")
        actions = _strip_markdown_bullets(analysis.get("doctor_summary", ""))
        ws.append([patient_id, med_summary, lab_summary, actions])

    # ── Styling ───────────────────────────────────────────────────────────────

    @staticmethod
    def _apply_styles(ws: Any, sheet_name: str) -> None:
        widths = COL_WIDTHS.get(sheet_name, [])
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill = ALT_FILL if row_idx % 2 == 0 else None
            for cell in row:
                cell.font = BODY_FONT
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if fill:
                    cell.fill = fill

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 22


def _strip_markdown_bullets(text: str) -> str:
    """Replace Unicode bullet (•) with '-' for plain Excel cells."""
    text = text.replace("\u2022", "-")
    text = re.sub(r"\*\*(.+?)\*\*", r"\1", text)  # remove **bold** markers
    return text
